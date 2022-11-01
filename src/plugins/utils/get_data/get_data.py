from typing import List

types = {
    "csv":"get_data_from_csv",
    "txt":'get_data_from_txt',
    "excel":"get_data_from_excel",
    "json":"get_data_from_json"
}

#Import cvs file from path
def get_data_from_csv(path_file:str, delimiter:str = ",", quotechar:str = "|"):
    import csv
    with open(path_file, newline='') as csvfile:
        data = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
        
    return data
#Import text file from path
def get_data_from_txt(path_file:str, encoding:str = 'utf-8'):
    try:
        with open(path_file, "r", encoding = encoding) as textfile:
            data = [[el.strip() for el in line.split(",")] for line in textfile.readlines()]
        return data
    except: 
        raise
#import json file from path
def get_data_from_json(path_file:str):
    import json
    try:
        with open(path_file, "r") as jsonfile:
            data = json.loads(jsonfile)
        return data 
    except:
        raise
#import excel from path   
def get_data_from_excel(path_file:str):
    import openpyxl
    wookbook = openpyxl.load_workbook(path_file)
    worksheet = wookbook.active
    data = []
    for i in range(0, worksheet.max_row):
        row = []
        for col in worksheet.iter_cols(1, worksheet.max_column):
            row.append(col[i])
        data.append(row)
    return data

class ExtensionTypeNotSupportedException(Exception):
    def __init__(self, extension) -> None:
        self.extension = extension
        self.messege = f"Files Extensions *.{extension} is not implemented"
        super().__init__(self.messege)

#receiving path to apply a function to import     
def get_data(path_file,**kwargs):
    extension = path_file.split(".")[-1]
    if extension not in types.keys():
        raise ExtensionTypeNotSupportedException(extension)    
    
    return globals()[types[extension]](path_file, **kwargs)